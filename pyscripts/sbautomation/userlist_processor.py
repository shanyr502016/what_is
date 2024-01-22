import openpyxl
print("Hello World")
wb = openpyxl.load_workbook("C:\Work\Scratch\MakeUserExcel\TCSMOQ10_Users.xlsx")
make_user_sheet = wb["Make_User"]
filecontents = ""
for x in range(2, make_user_sheet.max_row):
  filerowcontents =""
  for y in range(1, 28):
    headervalue = (make_user_sheet.cell(row = 1, column = y).value)
    cellvalue = (make_user_sheet.cell(row = x, column = y).value)
    print(str(headervalue))
    print(str(cellvalue))
    if (str(cellvalue) != 'None'):
        if(y <= 6):
          # filerowcontents += "|"
           filerowcontents += str(cellvalue)
           filerowcontents += "|"
           print(filerowcontents)
        elif(y != 28):
           filerowcontents += str(headervalue)
           filerowcontents += "|"
           filerowcontents += str(cellvalue)
           filerowcontents += "|"
           print(filerowcontents)
        elif(y == 28):
           filerowcontents += str(headervalue)
           filerowcontents += "|"
           
  filecontents += filerowcontents
  filecontents += "\n"
  print(filerowcontents)
print("============Final File")
print(filecontents)
file = open("C:\Work\Scratch\MakeUserExcel\TCSMOQ10.txt", "w")
file.write(str(filecontents))
file.close
