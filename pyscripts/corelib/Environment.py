from corelib.Loggable import Loggable
from corelib.DateTime import DateTime
from corelib.KeePass import KeePass
import json
import os
import socket
import re
import platform
import time
import datetime
import sys
from corelib.SCP import SCP
from corelib.Process import Process
import tempfile
from openpyxl import load_workbook




WINDOWS = 1
LINUX = 2
SOLARIS = 3
CURRENT_PLATFORM = -1

class Environment(Loggable):

    def __init__(self):

        super().__init__(__name__)

        self._environment = {}

        global CURRENT_PLATFORM
        if CURRENT_PLATFORM != -1:
            return
        uname = platform.system()
        if 'Windows' in uname:
            CURRENT_PLATFORM = WINDOWS
        elif 'SunOS' in uname:
            CURRENT_PLATFORM = SOLARIS
        else:
            CURRENT_PLATFORM = LINUX

       

    def get_environment_name(self):

        """
        Returns environment name comes from launcher command
        """
        return os.getenv('ENVIRONMENT_NAME')
    

    def get_infodba_credentials(self, server_info):

        args = []
        if 'TC_USER' in server_info and server_info['TC_USER'].__ne__(''):
            args.extend(['-u='+server_info['TC_USER']])
        else:
            self.log.error('TC_USER is required')
            # exit()
        if 'TC_PWF' in server_info and server_info['TC_PWF'].__ne__(''):
            args.extend(['-pf='+server_info['TC_PWF']])
        else:
            self.log.error('TC_PWF is required')
            # exit()
        return args
    
    def get_group(self, server_info):

        args = []
        if 'TC_GROUP' in server_info and server_info['TC_GROUP'].__ne__(''):
            args.extend(['-g='+server_info['TC_GROUP']])
        else:
            self.log.error('TC_GROUP is required')
            # exit()
        return args
    
    def get_project_credentials(self, server_info):
        args = []
        if 'PROJECT_ADMIN' in server_info and server_info['PROJECT_ADMIN'].__ne__(''):
            args.extend(['-u='+server_info['PROJECT_ADMIN']])
        else:
            self.log.error('PROJECT_ADMIN is required')

        if 'PROJECT_PWF' in server_info and server_info['PROJECT_PWF'].__ne__(''):
            args.extend(['-pf='+server_info['PROJECT_PWF']])
        else:
            self.log.error('PROJECT_PWF is required')
        return args

    def get_project_group(self, server_info):
        args = []
        if 'PROJECT_GROUP' in server_info and server_info['PROJECT_GROUP'].__ne__(''):
            args.extend(['-g='+ '"' + server_info['PROJECT_GROUP'] + '"'])
        else:
            self.log.error('PROJECT_GROUP is required')
        return args

    
    def get_property_validation(self, keyname, keyObject):

        if keyname in keyObject and keyObject[keyname].__ne__(''):
            return True
        else:            
            self.log.error(f"{keyname} is missing. Please check the configuration")
            # exit()

    def wait_execute(self, seconds, printmessage=True):
        self.log.info("Please wait.....")
        wait = seconds
        while wait > 0:
            if printmessage:
                sys.stdout.write('-')
            time.sleep(1)
            sys.stdout.flush()
            wait = wait - 1
        if printmessage:
            sys.stdout.write('\n')
    
    def get_share_root_path(self):
        global CURRENT_PLATFORM
        """
        Returns smo share location root path
        """
        if CURRENT_PLATFORM == LINUX:
            return os.environ.get('SMO_SHARE_LNX_ROOT')
        if CURRENT_PLATFORM == WINDOWS:
            return os.environ.get('SMO_SHARE_WIN_ROOT') 
        
    def get_share_root_us_path(self):
        global CURRENT_PLATFORM
        """
        Returns smo share location root path
        """
        if CURRENT_PLATFORM == LINUX:
            return os.environ.get('SMO_SHARE_LNX_ROOT_US')
        if CURRENT_PLATFORM == WINDOWS:
            return os.environ.get('SMO_SHARE_WIN_ROOT_US')

    def get_share_root_sg_path(self):
        global CURRENT_PLATFORM
        """
        Returns share location root path
        """
        if CURRENT_PLATFORM == LINUX:
            return os.environ.get('SMO_SHARE_LNX_ROOT_SG')
        if CURRENT_PLATFORM == WINDOWS:
            return os.environ.get('SMO_SHARE_WIN_ROOT_SG')

    def get_share_root_win_path(self):
        return os.environ.get('SMO_SHARE_WIN_ROOT')  

    def get_share_root_lnx_path(self):
        return os.environ.get('SMO_SHARE_WIN_ROOT')         
    
    def get_share_root_win_us_path(self):
        return os.environ.get('SMO_SHARE_WIN_ROOT_US')    
    
    def get_share_root_win_sg_path(self):
        return os.environ.get('SMO_SHARE_WIN_ROOT_SG')  

    def get_share_deploy_path(self):
        """
        Returns smo share location with deploy path
        """
        return os.getenv('SMO_PACKAGE_SHARE')
    
    def get_share_deploy_us_path(self):
        """
        Returns smo share location with deploy path
        """
        return os.getenv('SMO_PACKAGE_SHARE_US')
    
    def get_share_deploy_sg_path(self):
        """
        Returns smo share location with deploy path
        """
        return os.getenv('SMO_PACKAGE_SHARE_SG')

    def get_current_os_type(self):
        global CURRENT_PLATFORM
        if CURRENT_PLATFORM == LINUX:
            return 'LNX'
        if CURRENT_PLATFORM == WINDOWS:
            return 'WIN' 

    def get_tc_data(self):

        return os.environ.get('TC_DATA')

    def get_tc_root(self):

        return os.environ.get('TC_ROOT')
    
    def get_tc_temp(self):
        self.log.debug(f"TC TEMP Directory: {os.environ.get('TC_TMP_DIR')}")
        return os.environ.get('TC_TMP_DIR')
    
    def get_dita_log_path(self):

        return os.getenv('DITA_LOG_LOCATION')
    
    def get_tc_package_info(self, tcpackage_id):
        tc_package_name = []
        for tc_package in tcpackage_id.split(','):
            tc_package_name.append(tc_package[:-13])
        tc_package_info = '__'.join([str(elem) for elem in tc_package_name]) + tcpackage_id[-13:]
        return tc_package_info
    
    def change_execmod(self, path, permission='0777'):
        
        self.log.debug(f"File Permission Changed: {permission} {path} ")
        os.system(f'chmod -R {permission} {path}' )


    def get_base_path(self):
        return os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
    
    def get_instances(self, modulename):
        
        instance = self._environment['properties'][modulename]['NODE']
        
        self.log.info(self._environment['properties'][instance])
        
        return instance

    def get_ssh_command(self, nodes, removeSSH=False): 
        
        if 'USERNAME' and 'HOSTNAME' in nodes:
            if self.get_hostname() == nodes['HOSTNAME']:
                return ''
            else:
                _ssh_command = []   
                if not removeSSH:
                    _ssh_command.append('ssh')
                _ssh_command.append(nodes['USERNAME'] + '@'+ nodes['HOSTNAME'])
                _ssh_command = " ".join(_ssh_command)
                return _ssh_command
        else:
            return ''
        
    def is_linux(self, type):
        
        if type == 'LNX':
            return True
        else:
            return False

        
    def is_windows(self, type):
        
        if type == 'WIN':
            return True
        else:
            return False    
        
    def get_servers_file(self):
        
        return os.path.join(os.path.abspath(os.path.join(self.get_base_path(), '..')), '.config','servers',self.get_environment_name()+'.environment.json')

    def get_environment_file(self, package_info):
        
        if package_info.__eq__('deployment'):            
            return os.path.join(os.path.abspath(os.path.join(self.get_base_path(), '..')), '.config','Deploy','deploy_targets.json')
        elif package_info.__eq__('build'):
            return os.path.join(os.path.abspath(os.path.join(self.get_base_path(), '..')), '.config','Build','build_targets.json')
        else:
            return os.path.join(self.get_base_path(), 'config', self.get_environment_name() + '.config.json')

    def get_security_file(self, package_info):
    
        self.log.info(self.get_environment_name())
        
        return os.path.join(os.path.abspath(os.path.join(self.get_base_path(), '..')), '.config','Security',self.get_environment_name()+'.security.json')
    
    def get_dcscript_file(self):

        return os.path.join(os.path.abspath(os.path.join(self.get_base_path(), '..')), '.config', 'DCScript', f'{self.get_environment_name()}.dcscript.json')


    def read_system_properties(self, package_info,  printservice: bool, callback_count: int):        
        try:
            with open(self.get_environment_file(package_info)) as file:
                data = json.load(file)
                if not printservice and int(callback_count) == 1: 
                    self.log.info(f"Read Environment configuration from {str(self.get_environment_name())}")
                if package_info.__ne__('deployment') or package_info.__ne__('build'):
                    for key,val in data.items():
                        if data[key] != self.get_environment_name() and 'INSTANCES' in data[key]:
                            if all(isinstance(item, str) for item in data[key]['INSTANCES']):
                                sub_data = [data[_sub] for _sub in data[key]['INSTANCES'].split(',')]
                                sub_data = [instance for d in sub_data for instance in d['INSTANCES']]
                                ins_data = {'INSTANCES': sub_data}
                                data[key] = ins_data
                if 'MANAGE' in data:
                    self._environment['manage_services'] = data['MANAGE']
                    data.pop('MANAGE')
                if 'START' in data:
                    self._environment['start_services'] = data['START']
                    data.pop('START')
                if 'STOP' in data:
                    self._environment['stop_services'] = data['STOP']
                    data.pop('STOP')
                self._environment['properties'] = data
                return data
        except Exception as exp:
            self.log.error(f"Configuration Error: {exp}")
            self._environment['properties'] = {}
            return {}
        
    def read_environment_properties(self, package_info, printservice: bool, callback_count: int):
        """
        Read environment based configuration from .config/servers/{environment}.environment.json
        """
        with open(self.get_servers_file()) as file:
            data = json.load(file)
            if not printservice and int(callback_count) == 1: 
                self.log.info(f"Read Servers configuration from {str(self.get_environment_name())}")
            for node, server_info_list in data.items():
                if isinstance(server_info_list,list):
                    for server_info in server_info_list:
                        if isinstance(server_info,dict):
                            server_info['NODE'] = node
            self._environment['environment'] = data
            return data

    def read_security_properties(self, package_info, printservice: bool, callback_count: int):
        """
        Read security information from .config/Security/{environment}.security.json
        """
        with open(self.get_security_file(package_info)) as file:
            data = json.load(file)
            #if not printservice and int(callback_count) == 1: 
            #    self.log.info(f"Read Security configuration from {str(self.get_environment_name())}")
            self._environment['security'] = data
            return data 

    def read_security_keepass(self, package_info, printservice: bool, callback_count: int):
        try:
            self._keepass = KeePass()
            _encrypted_keepass_pass = 'UbUF1msD/mNr8+i7QFZdXGKL/C01fky1Td5+aVlcBKG0UUlT2/PW6rHvco6xuJWE'
            data = self._keepass.get_keepass("/smo_share/dita_share_do_not_modify/keepass/SMOTCDEV.kdbx", _encrypted_keepass_pass, self.get_environment_name()) 
            if not printservice and int(callback_count) == 1:
                self.log.info(f"Read Security configuration from {str(self.get_environment_name())}")
            self._environment['keepass'] = data
        except Exception as exp:
            self.log.error(f'Read KeePass failed')
        
    def read_properties_file(self, filepath):
        """
        Read Properties file and make object return
        """
        properties = dict()
        try:
            with open(filepath, 'r') as filecontent:
                for line in filecontent:
                    if '=' in line:
                        name, value = line.split('=', maxsplit=1)
                    properties[name.strip()] = value.strip()
            return properties
        except Exception as exp:
            self.log.error(f'Read properties {filepath} failed')
            
    def read_excel_sheet(self, excel_sheet_path, sheet_name):
        try:
            # Load the workbook (Excel file)
            workbook = load_workbook(excel_sheet_path)
            
            # Get the sheet by name or index (Sheet1 in this case)
            sheet = workbook[sheet_name]  # or by index: workbook.worksheets[0]
            
            # Initialize an empty list to store the data
            data = []
            # Get header row as keys
            keys = [str(cell.value).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

            # Iterate through rows (excluding the header row) and construct dictionaries
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # Start from row 2 (excluding header)        
                # Add row number as the first element in the row
                # Create a modified keys list with the custom key
                keys_with_custom = ['row_index'] + keys           
                # Trim whitespace from cell values and handle datetime.time(0, 0) as None
                row_data = [str(cell).strip() if cell is not None else None for cell in row]
                row_dict = dict(zip(keys_with_custom, [row_index] + row_data))
                data.append(row_dict)
            
            return data
        except FileNotFoundError:
            self.log.error("The specified file was not found.")
        except openpyxl.utils.exceptions.InvalidFileException:
            self.log.error("The specified file is not a valid Excel file.")
        except Exception as e:
            self.log.error(f"An error occurred: {str(e)}")
            
    def write_excel_sheet(self, excel_sheet_path, sheet_name, column_name, row_index, cell_value):
        try:
            # Load the workbook (Excel file)
            workbook = load_workbook(excel_sheet_path)
            
            # Get the sheet by name or index (Sheet1 in this case)
            sheet = workbook[sheet_name]  # or by index: workbook.worksheets[0]
            
            # Find the column index based on the column name in the first row
            column_index = None
            for col_idx, cell in enumerate(sheet[1], start=1):  # Assume the column names are in the first row (index 1)
                if cell.value == column_name:
                    column_index = col_idx
                    break
                    
            if column_index is not None:
                # Update cell value at row_index and column_index
                cell = sheet.cell(row=row_index, column=column_index)
                cell.value = cell_value  # Update with the desired value
            else:
                self.log.error(f"Column '{column_name}' not found in the first row of '{sheet_name}'.")                
            # Save the workbook after making modifications
            workbook.save(excel_sheet_path)
            return True
        except FileNotFoundError:
            self.log.error("The specified file was not found.")
            return False
        except openpyxl.utils.exceptions.InvalidFileException:
            self.log.error("The specified file is not a valid Excel file.")
            return False
        except Exception as e:
            self.log.error(f"An error occurred: {str(e)}")
            return False
    
    def get_build_package_name(self, software_version, package_type, branch_name, package_name):

        return software_version + '_' + package_type + '_' + branch_name + '_' + package_name   
        

    def get_os_environ(self):
        """
        Returns the environment paths
        """

        return os.environ
    
    def get_hostname(self):
        """
        Returns the name of this host
        """        
        return socket.gethostname()  
    
    def get_execute_targets(self, module_name):
        """
        Returns multiple execute targets from Properties configuration
        """        
        return self._environment['properties'][module_name]['executeTargets']
    
    def get_parallel_status(self, module_name):
        """
        Returns multiple executetargets parallel status from Properties configuration
        """  
        if 'parallel' in self._environment['properties'][module_name]:      
            return self._environment['properties'][module_name]['parallel']
        return False
    
    def get_execute_target(self, module_name, sub_module_name):
        """
        Returns single execute targets from Properties configuration
        """        
        return  module_name + '.' + '.'.join([str(elem) for elem in sub_module_name])
    
    def get_location_in_package(self, package_id, location = ''):
        """
        Get Package location from Dev Share Location or Prod Share Location
        """

        package_id = package_id.strip()

        for release_package_folder_name in os.listdir(os.path.join(self.get_share_deploy_path(), 'SMO')):

            if os.path.isdir(os.path.join(self.get_share_deploy_path(), 'SMO', release_package_folder_name)):             

                if re.search(release_package_folder_name,package_id):
                        
                    for release_package in os.listdir(os.path.join(self.get_share_deploy_path(),'SMO', release_package_folder_name)):                        

                        if re.search(release_package,package_id):

                            for release_package_id in os.listdir(os.path.join(self.get_share_deploy_path(),'SMO', release_package_folder_name, release_package)): 

                                try: 
                                    if release_package_id == package_id:  
                                        return os.path.join(self.get_share_deploy_path(), 'SMO', release_package_folder_name, release_package, release_package_id, location)
                                except Exception:
                                    self.log.info(package_id)
                                    self.log.error('Package not Found')  
        
        self.log.warning('Package not Found')
        return ''                     

        # for release_package_folder_name in os.listdir(self.get_share_deploy_path()):
        #     if os.path.isdir(os.path.join(self.get_share_deploy_path(), release_package_folder_name)):                
        #         for release_package_id in os.listdir(os.path.join(self.get_share_deploy_path(),release_package_folder_name)):
        #             try:
        #                 if release_package_id == package_id:                        
        #                     return os.path.join(self.get_share_deploy_path(), release_package_folder_name, release_package_id, location)
        #             except Exception:
        #                 self.log.error('Package not Found')
                        
                        
    def get_backup_location(self, targetpath, package_id, hostname):
        return os.path.join(targetpath, package_id + '_' + DateTime.get_datetime("%d%m%Y"),hostname.lower())  
    
    def get_execute_server_details(self, execute_server):

        try:
            if execute_server in self._environment['environment']:
                if 'executeServers' in self._environment['environment'][execute_server]:
                    executeServers = self._environment['environment'][execute_server]['executeServers']
                    executeServers = executeServers.split(',')
                    execute_servers = []
                    for index, executeServer in enumerate(executeServers):
                        if executeServer in self._environment['environment']:
                            execute_servers.append(self._environment['environment'][executeServer][0])
                        else:
                            self.log.warning(f'{executeServer} Server not available in this {self.get_environment_name()} Environment')                        
                    return execute_servers
                else:
                    return self._environment['environment'][execute_server]
            else:
                self.log.warning(f'{execute_server} Server not available in this {self.get_environment_name()} Environment. Skip the warning')
                return []
        except Exception as exp:
            self.log.error(f'ExecuteServer not exists! {exp}')      
        
    

    def getTCPackageID(self, tc_package_id):
        tc_packages = []
        def getPackage(package_name):
            for tc_id in tc_package_id.split(','):            
                if re.search(package_name,tc_id):
                    return tc_id
        
        CMNPackage = getPackage(r'_CMN')
        if CMNPackage:
            tc_packages.append(CMNPackage)
        
        RIPackage = getPackage(r'_RI')
        if RIPackage:
            tc_packages.append(RIPackage)
        return tc_packages
    
    def get_current_execution_server(self):
        """
        Get Current execution based on HOSTNAME self.get_environment_name()
        """
        try:
            for node,server_info_list in self._environment['environment'].items():
                if isinstance(server_info_list,list):
                    for server_info in server_info_list:
                        if isinstance(server_info,dict):
                            if server_info['HOSTNAME'].lower() == self.get_environment_name().lower():
                                return node
            return '-'
        except Exception as exp:
            self.log.error(f'ExecuteServer not exists! {exp}')


    def clear_temp_dirs(self, location_list:list, server_info):
        try:
            _result = 1
            self.log.info('Temporary Directories Delete - Started')
            _scp_without_ssh = SCP(self.get_ssh_command(server_info, True)) 
            _scp_with_ssh = SCP(self.get_ssh_command(server_info)) 
            __ssh_command = self.get_ssh_command(server_info)
            _is_windows = self.is_windows(server_info['OS_TYPE'])
            lines = []
            for temp_dir in location_list:
                temp_dir = temp_dir.replace('$USERNAME',server_info['USERNAME'])
                check_path_exists = _scp_with_ssh.check_dir_to_remote(_is_windows,temp_dir)
                if check_path_exists:
                    lines.append(f'Remove-Item -Path "{temp_dir}" -Recurse -Force -ErrorAction SilentlyContinue -ErrorVariable +Errors;if ($Errors) {{ Write-Host {temp_dir} }}')
                else:
                    self.log.warning(f'{temp_dir} Directory does not exists to delete')
            if len(lines):
                file  = tempfile.NamedTemporaryFile('w+t')
                filepath = file.name
                for line in lines:
                    self.log.debug(line)
                    file.write(line + '\n')
                file.flush()
                resultshare = _scp_without_ssh.copy_to_remote(filepath, os.path.join(server_info['DEPLOYMENT_CENTER_TEMP_DIR'],'removetempfiles.ps1'))
                if resultshare == 0:
                    args = [__ssh_command]
                    args.append('"cd '+server_info['DEPLOYMENT_CENTER_TEMP_DIR']+' && '+server_info['DEPLOYMENT_CENTER_TEMP_DIR'].split(':')[0]+':'+' && powershell '+server_info['DEPLOYMENT_CENTER_TEMP_DIR']+'/removetempfiles.ps1"')
                    args = ' '.join([str(elem) for elem in args])
                    process = Process(args)
                    process.hide_output()
                    process.collect_output()
                    process.ignore_errors()
                    process.execute()
                    output_result = [res for res in process.get_out_lines() if res != '']
                    if len(output_result):
                        for undeleted_path in output_result:
                            self.log.warning(f'{undeleted_path} is Not deleted...! May be the file or Directory is in Use')
                    _result = 0
                else:
                    _result = 1
            else:
                _result = 0
                self.log.warning(f'Directories does not exists to delete')
        except Exception as exp:
            self.log.error(exp)
            _result = 1
        return _result


    


    def resume_state_log_path(self,tcpackage_id,location = ''):
        return os.path.join(Loggable.get_config_provider().get_dita_log_path(),self.get_tc_package_info(tcpackage_id).replace(',','__'),location)     

    def resume_state_calculation_result(self,processlist:list):
        """
        Resume state calculation based on the processlist coming from exection targets results.
        """

        _result = {}
        for _process in processlist:
            server_name = _process['NODE']
            if server_name not in _result:
                _result[server_name] = []
            _result[server_name].append(_process)

        _processlist = []
        for server, items in _result.items():
            if len(items) > 1:
                for item in items:
                    if item['process'] != 0:
                        _processlist.append(item)
                        break
                else:
                    _processlist.append(items[-1])
            else:
                _processlist.append(items[0])
        return _processlist


    def get_target_with_server(self,targetList:list, server_list:list,new_state = True):

        result = {}
        for command, server in zip(targetList, server_list):
            if command not in result:
                result[command] = [server]
            else:
                result[command].append(server)

        if new_state:
            return {key:val[0] for key,val in result.items()}

        return result

    def update_serverinfo_data(self,targets:list,arguments):
        """
        update server info if executionServer key not available
        """
        result = []
        try:
            for i in targets:
                if len(i):
                    result.extend(arguments._environment[i])
            return result
        except Exception as exp:
            self.log.error(f'Resume State Read Failed (update_serverinfo_data)- {exp}')
            return result

    def resume_state_execution_read(self,tc_package_id:str,module_name:str,resume:bool,single_target = False):
        """
        Resume state exection reading the txt file, any failed status is avaliable next execution it will resume the state if you use same package name.
        If don't want this feature, add -r in the dita command
        """

        try:

            log_path = self.resume_state_log_path(tc_package_id)
            filename = os.path.join(log_path,f'{module_name}_{tc_package_id}.txt')                

            target_list = []
            if 'executionServer' in self._environment['properties'][module_name]:
                target_list = [module_name]
            else:
                for target in self._environment['properties'][module_name]['executeTargets'].split(','):
                    if 'executeTargets'  in self._environment['properties'][target]:
                        targets = [target for target in self._environment['properties'][target]['executeTargets'].split(',')]
                    else :
                        targets = [target]
                    target_list.extend(targets)

            _targetList = []
            for _target in target_list:
                if 'executionServer' in self._environment['properties'][_target]:
                    for _index,_server_info in enumerate(self.get_execute_server_details(self._environment['properties'][_target]['executionServer'])):
                        _targetList.append(_target+'-'+_server_info['NODE']+ '-' +'FAILED\n')
                else:
                    _targetList.append(_target+'-'+ self.get_current_execution_server() + '-' +'FAILED\n')

            content_list = [con for con in _targetList if 'FAILED' in con]
            content_list = [content.split('-')[:2] for content in  content_list]
            output_dict = {key: [item[key] for item in [{item[0]:item[1]} for item in content_list] if key in item] for d in [{item[0]:item[1]} for item in content_list] for key in d}

            def update_execution_target_data(filename:str,target_list:list):    
                with open(filename,'w') as _errorTargetData:
                    for _target in target_list:
                        if 'executionServer' in self._environment['properties'][_target]:
                            for _index,_server_info in enumerate(self.get_execute_server_details(self._environment['properties'][_target]['executionServer'])):
                                _errorTargetData.writelines(_target+'-'+_server_info['NODE']+ '-' +'FAILED\n')
                        else:
                            _errorTargetData.writelines(_target+'-'+ self.get_current_execution_server() + '-' +'FAILED\n')

            if not resume:

                with open(filename,'a+') as errortargets:
                    errortargets.seek(0)
                    content_list = [con for con in errortargets.readlines() if 'FAILED' in con]

                if len(content_list):
                    _targetList =[content.split('-')[:2] for content in  content_list]
                    output_dict = {key: [item[key] for item in [{item[0]:item[1]} for item in _targetList] if key in item] for d in [{item[0]:item[1]} for item in _targetList] for key in d}

                else:
                    update_execution_target_data(filename,target_list)

            else:
                update_execution_target_data(filename,target_list)

            return output_dict

        except Exception as exp:
            self.log.debug(f'Resume State Read Failed - {exp}')
            return {}



    def resume_state_execution_write(self,tc_package_id:str,module_name:str,target_name:str,processlist:list,parellel = False):
        """
        Resume state the Deployment incase any failure happens in Multiple targets combined, write the success and failure state into the txt file. (PreBMIDE, PostBMIDE, Backup, Restore, etc..)
        """
        try:

            log_path = self.resume_state_log_path(tc_package_id)
            filename = os.path.join(log_path,f'{module_name}_{tc_package_id}.txt')

            if len(processlist):
                _processlist = self.resume_state_calculation_result(processlist)

            else:
                _processlist = processlist

            _result_content = []
            _result = 1
            for _index,_process in enumerate(_processlist):
                if isinstance(_process['process'],int):
                    _result= _process['process']
                else:
                    _result= _process['process'].wait()
                if _result == 0:
                    content = target_name+'-'+ _process['NODE'] + '-' +'SUCCESS\n'
                else:
                    content = target_name+'-'+ _process['NODE'] + '-' +'FAILED\n'

                _result_content.append(content)

            new_results_dict = {result.rsplit('-', 1)[0]: result.rsplit('-', 1)[1] for result in _result_content}

            def replace_status(match, new_results):
                module_server = match.group(1)
                new_result = new_results.get(module_server)
                return f"{module_server}-{new_result}" if new_result else match.group(0)

            with open(filename,'r') as _existing_content:
                existing_data = _existing_content.readlines() 

            updated_data = [re.sub(r'^(.*?-\w+)-(FAILED|SUCCESS)\n$', lambda match: replace_status(match, new_results_dict), data) for data in existing_data]

            with open(filename,'w') as _update_content:
                _update_content.writelines(updated_data)

            return _result

        except Exception as exp:
            self.log.error(f'Resume State Write Failed - {exp}')
            return 1

    def read_delta_targets(self,targets:str):
        try:
            _result = []
            target_list= [target.lower() for target in targets.strip().split(',')]
            for _properties in self._environment['properties']:
                if _properties.lower() in target_list:
                    _result.append(_properties)
            if len(_result) :
                if len(_result) != len(target_list):
                    self.log.warning(f'Some delta targets mismatched - Check the Delta targets...')
                return ','.join(_result)
            else:
                self.log.warning('Given delta targets are not found...Executing the targets From Configuration...')
                return False

        except Exception as exp:
            self.log.error(f'Delta Execution Failed - {exp}')
            return False
            

